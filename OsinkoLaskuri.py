import openpyxl
from openpyxl import Workbook

wb = Workbook()

## ----- VARIABLES FOR USER TO CHOOSE ----- ##

#Stocks that you don't want to include. For example non-divident stocks.
BlackList = ['OP-AASIA INDEKSI A','OP-AMERIKKA INDEKSI A','OP-EUROOPPA INDEKSI A']
BlackListEnabled = True
#All possible foreign currensies. 
ForeignCurrencies = ['SEK', 'USD']
#File names
OpenFileName = "Sijoitukset.xlsx"
SaveFileName = "Sijoitukset2.xlsx"

## ----- NON USER VARIABLES ----- ##
Events = []

def ReadInvestments():
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook(OpenFileName)
     
    # Define variable to read sheet
    worksheet1 = dataframe.active
     
    # Iterate the loop to read the cell values
    for row in range(1, worksheet1.max_row ): #worksheet1.max_row
        L = []
        for col in worksheet1.iter_cols(0, 7):
            data = col[row].value
            if data == None:
                L.append('0.0')
            else:
                L.append(str(data))
        
        if L[1] == "OSTO":       
            L.append("TOSI")
            L.append(0) #Osinko (€) sarake
            L.append(0) #Osinko (%) sarake
        else: #Muihinkuin ostoihin tyhjiä soluja
            L.append(None)
            L.append(None)
            L.append(None)          

        #Blacklist
        if BlackListEnabled == True:
            if L[0] not in BlackList:
                Events.append(L)
        else:
            Events.append(L)
    print("# ---------- Read Investments: Done! ---------- #")
      

def ClearBadData():
    for i in range(len(Events)):
        
        Splitted = Events[i][2].split(".")
        Splitted.reverse()
        Events[i][2] = Splitted[0] + "." + Splitted[1] + "." +  Splitted[2]
        
        E = Events[i][3]
        E = E.replace("kpl","")
        E = E.replace(",",".")
        Events[i][3] = float("".join(E.split())) #Kept as float, because indexes amounts can be desimal numbers
    
        
        E = Events[i][5]
        E = E.replace("EUR","")
        E = E.replace(",",".")
        Events[i][5] = float("".join(E.split()))
        
        
        
        E = Events[i][6]
        E = E.replace("EUR","")
        E = E.replace(",",".")
        Events[i][6] = float("".join(E.split()))

        #Kurssi lasketaan hinta jaettuna osake

        Events[i][4] = (Events[i][6] - Events[i][5]) / Events[i][3] #Converts non-euros to euros
        
        
    print("# ---------- Clear Data: Done! ---------- #")
    
def myFunc(e):
  return e[0]

def Sort():
    global Events
    Events = sorted(Events, key=lambda x: (x[0], x[2]))
    print("# ---------- Sort: Done! ---------- #")
    
def DividePurchases():
    #Taxes
    for i in range(len(Events)):
        if Events[i][1] == "VERON PIDÄTYS":
            if Events[i-1][1] == "OSINKO" or Events[i][0] == Events[i-1][0]:
                
                PerOsake = float(Events[i][6]) / float(Events[i][3])
                Events[i-1][4] -= PerOsake
            else:
                print("ERROR at Events[" + str(i) + "]")
    ## ----- Splitting bought of stocks depending on sales ----- ###
    CompName = ""
    CompIndex = 0
    i = 0
    while i < len(Events):
        if Events[i][0] != CompName:
            CompName = Events[i][0]
            CompIndex = i
        if Events[i][1] == "MYYNTI":
            HoldIndex = CompIndex
            SoldAmount = Events[i][3]
            
            while SoldAmount > 0:
                if Events[HoldIndex][7] == "TOSI" and Events[HoldIndex][1] == "OSTO":
                    if SoldAmount == Events[HoldIndex][3]: #Eivät mahdu yhteen. 
                        Events[HoldIndex][7] = "EPÄTOSI"
                        SoldAmount -= Events[HoldIndex][3]
                        break
                    if SoldAmount > Events[HoldIndex][3]: #Eivät mahdu yhteen. 
                        Events[HoldIndex][7] = "EPÄTOSI"
                        SoldAmount -= Events[HoldIndex][3]
                        HoldIndex += 1
                        continue
                    if SoldAmount < Events[HoldIndex][3]: #Mahtuvat yhteen ostoon
                        #Holdindexsiin ne mitkä on myyty, lisätty on myymättömät
                        Events[HoldIndex][5] = float(Events[HoldIndex][5])
                        
                        NotSold = Events[HoldIndex][3] - SoldAmount
                        ProsentChange = float(SoldAmount / Events[HoldIndex][3])
                        TotalHolder = Events[HoldIndex][6]
                        ExchangeCost = Events[HoldIndex][5]
                        Events[HoldIndex][3] = SoldAmount
                        Events[HoldIndex][5] = ExchangeCost * ProsentChange
                        Events[HoldIndex][6] = ProsentChange * TotalHolder
                        Events[HoldIndex][7] = "EPÄTOSI"
                        
                        NewEventTotal = TotalHolder * (1 - ProsentChange)
                        E = Events[HoldIndex]
                        Events.insert(HoldIndex + 1,[E[0],E[1],E[2],NotSold, E[4], ExchangeCost * (1 - ProsentChange),NewEventTotal, "TOSI",0,0])
                        i += 1
                        break
                HoldIndex += 1
        i += 1 

    #Palauttaa omistus arvon todeksi kaikkiin, koska tarvitaan osingon laskuun
    for i in range(len(Events)):
        if Events[i][1] == "OSTO":
            Events[i][7] = "TOSI"  
   
    
    
def CalculateDividend():
    ## ----- Divident calculation ----- ##
    CompName = ""
    CompIndex = 0
    for i in range(len(Events)):
        if Events[i][0] != CompName:
            CompName = Events[i][0]
            CompIndex = i
            
        if Events[i][1] == "OSINKO":
            Amount = Events[i][3]
            Value = Events[i][4]
            
            HoldIndex = CompIndex
            while Amount > 0:
                if Events[HoldIndex][0] != CompName:
                    print("SOMETHING FUCKED UP! PLEASE CHECK THIS!" + CompName)
                    break
                if Events[HoldIndex][7] == "EPÄTOSI" or Events[HoldIndex][7] == None or Events[HoldIndex][1] != "OSTO":
                    HoldIndex += 1
                    continue
                if Events[HoldIndex][7] == "TOSI":
                    if Events[HoldIndex][3] < Amount:
                        Events[HoldIndex][8] += Events[HoldIndex][3] * Value
                        Amount -= Events[HoldIndex][3]
                        HoldIndex += 1
                        continue
                    if Events[HoldIndex][3] == Amount:
                        Events[HoldIndex][8] += Events[HoldIndex][3] * Value
                        break
                    if Events[HoldIndex][3] > Amount:
                        Events[HoldIndex][8] += Amount * Value
                        break
                else:
                    HoldIndex += 1 
        ## MYYNTI SUHTEESSA AIKAJANAAN
        if Events[i][1] == "MYYNTI":
            Amount = Events[i][3]
            HoldIndex = CompIndex
            while Amount > 0:
                if Events[HoldIndex][1] == "OSTO":
                    if Events[HoldIndex][7] == "TOSI":
                        #Should always either fit or go over
                        if Amount == Events[HoldIndex][3]:
                            Events[HoldIndex][7] = "EPÄTOSI"
                            break
                        if Amount > Events[HoldIndex][3]:
                            Events[HoldIndex][7] = "EPÄTOSI"
                            Amount -= Events[HoldIndex][3]
                HoldIndex += 1
    
    #Dividend procent
    for i in range(len(Events)):
        if(Events[i][1] == "OSTO"):
            Events[i][9] = Events[i][8] / Events[i][6]
    
    
    print("# ---------- Calculate Divident: Done! ---------- #")        

def SetData():
    worksheet2 = wb.active
    worksheet2.title = "Results"
    Titles = ['Laji:','Tapahtuma:','Päivämäärä:','Määrä:','Kurssi:','Kulut:','Yhteensä:','Omistus:','Osinko (€):','Osinko (%)']
    for i in range(len(Titles)):
        worksheet2.cell(row=1, column=i+1, value=Titles[i])
    skipped = 0
    for E in range(len(Events)):
        if Events[E][1] == "OSTO":
            for i in range(len(Events[E])):
                worksheet2.cell(row=E+2-skipped, column=i+1, value=Events[E][i])
        else:
            skipped += 1
    
    wb.save(SaveFileName)  
    print("# ---------- Set Data: Done! ---------- #")

def CheckType():
    print("Checking Types: ")
    for E in Events:
        print("Type: " + str(type(E[5])) + " Checking: " + str(E[4]))
    print("# ---------- Checkking Ended! ---------- #")

def printData():
    for i in range(len(Events)):
        print(Events[i])
    #print("# ---------- Printing Data Done! ---------- #")

def printDataRow(row = 0):
    print("#---------------------------- CHECKING DATA ROW START ---------------------------------- #")
    for i in range(len(Events)):
        print(Events[i][row])
    print("#-----------------------------CHECK DATA ROW END ------------------------------------#")
    
print("# ---------------------------------------------------------------------------------------------------- #")
ReadInvestments()
ClearBadData()
Sort()
printData()
DividePurchases()
CalculateDividend()
SetData()

print("YOUR DATA IS READY, SIR")